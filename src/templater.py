from config import ROOT_DIR, OUTPUT_PATH 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, Alignment, PatternFill, Border, Side
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict

def create_schema_sheet(
        lab:int | str, 
        question_array:List[str],
        schema_array:List[str],
        output_dir:Path) -> None:
    """Construct an xlsx schema template for a given lab.

    Construct and save an excel sheet a 'schema tab' containing the following
    columns:  
    - 'question' (numbering constructed by this function)
    - 'marks available' (left empty, fill in the excel sheet)
    - 'question summary' (left empty, fill in the excel sheet)
    - 'schema' (left empty, fill in the excel sheet)

    Args:
        lab(int): Lab number to generate template for.
        question_array (List[str]): A list containing the questions in that
        given lab.
        output_dir (Path): Output save path.

    Returns:
        Template schema dataframe for a given lab.
    """
    lab = "lab" + str(lab)
    questions = pd.DataFrame(
        {"question": [i for i in question_array]}
        )
    marks_available = pd.DataFrame(
        {"marks available":["" for n in question_array]}
        )
    summary = pd.DataFrame(
        {"question summary":["" for n in question_array]}
        )
    schema = pd.DataFrame(
        {"marking rubrick":["" for n in question_array]}
        )
    marks = pd.DataFrame(
        {"marks":["" for n in question_array]}
    )
    feedback = pd.DataFrame(
        {"feedback":["" for n in question_array]}
    )
    output = pd.concat(
        [questions, marks_available, summary, schema, marks, feedback], axis=1
        )
    save_path = output_dir / (lab + ".xlsx")
    output.to_excel(save_path, sheet_name="schema", index=False)

def create_grading_sheet(
        lab:int | str,
        output_dir:Path,
        question_array:List, 
        schema_array:List, 
        total_groups:int
        ) -> None:
    
    load_path = output_dir / ('lab' + str(lab) + ".xlsx")
    workbook = load_workbook(load_path)
    schema_sheet = workbook['schema'] #construct schema_array from schema_sheet
    try:
        grading_sheet = workbook['grading']
    except:
        grading_sheet = workbook.create_sheet("grading")
        grading_sheet = workbook['grading']

    def _interval_calc() -> Dict[str, List[str]]:
        """Returns a dict map of column header names as keys and a list of str
        representing columns for which that header must be applied.

        Essentially, which headers from the schema, when replicated (by as many
        times as there are questions) should be assigned to which column in the
        workbook. 

        Args:
            question_array (List): list of question numbers
            schema_array (List): list of schema headers to be repeated

        Returns:
            ...
        """
        repetitions = len(question_array)
        len_schema = len(schema_array)
        col_header_map = {}
        for i, schema_header in enumerate(schema_array):
            repeat_col_range = range(i+2, (len_schema*repetitions)+2, len_schema) 
            #i+2 because the first column of the grading sheet is reserved for a
            #'groups' column, so we skip it.
            col_header_map.update(
                {schema_header: [get_column_letter(x) for x in repeat_col_range]}
                )
        return col_header_map
    
    def _make_group_column() -> None:
        grading_sheet['A1'] = 'groups' #type: ignore
        for i in range(1, total_groups+1):
            #contents
            grading_sheet["A" + str((i+1))] = "group " + str((i)) # type: ignore

    def _make_repeat_columns(
            col_header_map:Dict[str, List[str]]) -> None:
        for header_name, target_columns in col_header_map.items():
            for idx, column in enumerate(target_columns):
                grading_sheet[column + "1"] = header_name # type: ignore
                if header_name == 'feedback' or header_name == 'mark':
                    continue
                else:
                    col_idx = column_index_from_string(column)
                    absolute_ref_col = get_column_letter(
                        column_index_from_string(target_columns[0]) - 1
                        )
                    for cells in grading_sheet.iter_cols( #type: ignore
                        min_col=col_idx, max_col=col_idx, min_row=2
                        ): 
                        for c in cells:
                            c.value = (
                                '=schema!$'+ absolute_ref_col +'$' + str(idx+2)
                                )
    _make_group_column()
    _make_repeat_columns(col_header_map=_interval_calc())
    workbook.save(load_path)

def style_sheets(lab:int | str, output_dir:Path) -> None:
    load_path = output_dir / ('lab' + str(lab) + ".xlsx")
    workbook = load_workbook(load_path)
    schema_sheet = workbook['schema']
    grading_sheet = workbook['grading']
    # HEADER TEXT:
    for c in grading_sheet['A']: # type: ignore
        c.font = Font(size=15, bold=True)
        c.alignment = Alignment(wrap_text=True, shrink_to_fit=False)
    for c in grading_sheet['1']: # type: ignore
        c.font = Font(size=15, bold=True)
        c.alignment = Alignment(wrap_text=True, shrink_to_fit=False)
    # COLUMN SPECIFIC / GRADING SHEET
    grading_sheet.column_dimensions["A"].width = 20 # type: ignore
    for column in grading_sheet.iter_cols(min_row=1, max_row=1): # type: ignore
        for cell in column:
            if cell.value == "question":
                grading_sheet.column_dimensions[cell.column_letter].width = 13 # type: ignore
            elif cell.value == "marks available":
                grading_sheet.column_dimensions[cell.column_letter].width = 13 # type: ignore
            elif cell.value == "question summary":
                grading_sheet.column_dimensions[cell.column_letter].width = 40 # type: ignore
                for c in grading_sheet[cell.column_letter]: # type: ignore
                    c.alignment = Alignment(wrap_text=True)
            elif cell.value == "marking rubrick":
                grading_sheet.column_dimensions[cell.column_letter].width = 40 # type: ignore
                for c in grading_sheet[cell.column_letter]: # type: ignore
                    c.alignment = Alignment(wrap_text=True)
            elif cell.value == "feedback":
                grading_sheet.column_dimensions[cell.column_letter].width = 40 # type: ignore
                for c in grading_sheet[cell.column_letter]: # type: ignore
                    c.alignment = Alignment(wrap_text=True)
    # SCHEMA SHEET
    schema_sheet.column_dimensions["A"].width = 15 # type: ignore
    schema_sheet.column_dimensions["B"].width = 40 # type: ignore
    schema_sheet.column_dimensions["C"].width = 40 # type: ignore
    schema_sheet.column_dimensions["D"].width = 40 # type: ignore
    schema_sheet.column_dimensions["E"].hidden = True # type: ignore
    schema_sheet.column_dimensions["F"].hidden = True # type: ignore
    for column in schema_sheet.iter_rows(): # type: ignore
        for c in column:
            c.font = Font(size= 15)
            c.alignment = Alignment(wrap_text=True)
    # ALTERNATING COLORS
    fill_light = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
    fill_dark = PatternFill(
        start_color="A9A9A9", end_color="A9A9A9", fill_type="solid"
        )
    for i, row in enumerate(grading_sheet.iter_rows(min_row=2), start=2): # type: ignore
        fill = fill_light if i % 2 == 0 else fill_dark
        for cell in row:
            cell.fill = fill
    #BORDERS
    thin_border_side = Side(border_style="thick", color="000000")
    border_right = Border(right=thin_border_side)
    for column in grading_sheet.iter_cols(min_row=1, max_row=1): # type: ignore
        for cell in column:
            if cell.value == 'feedback' or cell.value == 'groups':
                for row in grading_sheet.iter_rows(
                    min_row=2, min_col=cell.column, max_col=cell.column
                    ): # type: ignore
                    for cell in row:
                        cell.border = border_right
    workbook.save(load_path)

if __name__ == "__main__":
    METADATA_PATH = ROOT_DIR / "src" / "metadata.json"
    lab = input("Which lab number do you want to generate a template for? ")
    with open(METADATA_PATH) as f:
        metadata = json.load(f)
        # lab_assignment_relationships = metadata["assignments"]
        total_groups = metadata["totalGroups"] #type: int
        question_array = metadata["questionArray"]["lab" + lab] #type: List[str]
        schema_array = metadata["schemaArray"] #type: List[str]

    create_schema_sheet(
        lab=lab, 
        question_array=question_array, 
        schema_array=schema_array, 
        output_dir=OUTPUT_PATH
        )
    create_grading_sheet(
        lab=lab, 
        question_array=question_array, 
        schema_array=schema_array, 
        output_dir=OUTPUT_PATH, 
        total_groups=total_groups
        )
    style_sheets(
        lab=lab,
        output_dir=OUTPUT_PATH)
